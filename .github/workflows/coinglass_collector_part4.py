"""
Coinglass Data Collector - Part 4 of 4
Collects data for remaining pairs
"""

import asyncio
from playwright.async_api import async_playwright
import random
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime
import os
import time
from filelock import FileLock

EXCEL_PATH = os.path.join(os.getcwd(), "Trading_Journal.xlsx")
LOCK_FILE = os.path.join(os.getcwd(), "Trading_Journal.lock")

PAIRS = [
    "UNI", "HYPE", "TAO", "DOT", "NEAR", "FET", "BCH", "LTC", "LINK", "RENDER", "BNB"
]

URLS = {
    "open_interest": "https://www.coinglass.com/open-interest/{pair}",
    "currencies": "https://www.coinglass.com/currencies/{pair}",
    "volume": "https://www.coinglass.com/volume/{pair}",
}

COLUMNS = [
    "Date", "Time",
    "Long /short liquidation",
    "Liquidation Ratio",
    "Total Contracts",
    "OI Ch(24h)",
    "OI /24h_vol",
    "24 h vol (fut/spot)",
    "Price Performance 24h",
    "Long/short(24h) ratio (Binance/OKX/Bybit/Mexc)",
    "24 H vol",
    "7 day vs 24h vol",
    "30 day vs 24 h vol",
    "Net flow 24/3 day",
]

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 11.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
]

RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
BLUE_FILL = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
GREEN_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
RED_FONT = Font(color="CC0000", bold=True)
BLUE_FONT = Font(color="0000CC", bold=True)
GREEN_FONT = Font(color="006600", bold=True)


def clean_text(text):
    if not text:
        return ""
    return text.strip().replace("\n", " ").replace("\t", " ")


def parse_dollar_to_thousands(val):
    if not val:
        return 0
    val = str(val).strip().replace("$", "").replace(",", "")
    try:
        if "B" in val:
            return float(val.replace("B", "")) * 1000000
        elif "M" in val:
            return float(val.replace("M", "")) * 1000
        elif "K" in val:
            return float(val.replace("K", ""))
        else:
            return float(val) / 1000
    except:
        return 0


def calculate_liquidation_ratio(long_short_str):
    if not long_short_str or "/" not in long_short_str:
        return None
    try:
        parts = long_short_str.split("/")
        if len(parts) != 2:
            return None
        long_k = parse_dollar_to_thousands(parts[0].strip())
        short_k = parse_dollar_to_thousands(parts[1].strip())
        total = long_k + short_k
        if total == 0:
            return None
        ratio = (long_k - short_k) / total
        return round(ratio, 4)
    except Exception as e:
        print(f"  ⚠️ Liquidation ratio calc error: {e}")
        return None


def parse_to_million(val):
    if not val:
        return ""
    val = str(val).strip().replace("$", "").replace(",", "")
    try:
        if "B" in val:
            return round(float(val.replace("B", "")) * 1000, 2)
        elif "M" in val:
            return round(float(val.replace("M", "")), 2)
        elif "K" in val:
            return round(float(val.replace("K", "")) / 1000, 4)
        else:
            return round(float(val), 2)
    except:
        return val


def calc_vol_ratio(vol_24h, vol_period, days):
    try:
        v24 = float(vol_24h)
        vp = float(vol_period)
        daily_avg = vp / days
        ratio = v24 / daily_avg
        change = (v24 - daily_avg) / daily_avg * 100
        sign = "+" if change >= 0 else ""
        return ratio, change, f"{ratio:.4f} / {sign}{change:.2f}%"
    except:
        return None, None, ""


async def scrape_open_interest(page, pair):
    url = URLS["open_interest"].format(pair=pair)
    print(f"  OI: {url}")
    await page.goto(url, wait_until="domcontentloaded", timeout=30000)
    await asyncio.sleep(random.uniform(7, 12))

    data = {"total_contracts": "", "oi_ch_24h": "", "oi_24h_vol": ""}

    try:
        rows = await page.query_selector_all("tr")
        for row in rows:
            cells = await row.query_selector_all("td")
            if len(cells) >= 6:
                first_cell = clean_text(await cells[0].inner_text())
                if first_cell.strip() == "All":
                    if len(cells) >= 9:
                        data["total_contracts"] = clean_text(await cells[3].inner_text())
                        data["oi_ch_24h"] = clean_text(await cells[7].inner_text())
                        data["oi_24h_vol"] = clean_text(await cells[8].inner_text())
                    elif len(cells) >= 7:
                        data["total_contracts"] = clean_text(await cells[2].inner_text())
                        data["oi_ch_24h"] = clean_text(await cells[5].inner_text())
                        data["oi_24h_vol"] = clean_text(await cells[6].inner_text())
                    print(f"  ✅ OI: {data['oi_ch_24h']}")
                    break

        if not data["total_contracts"]:
            print("  Trying fallback...")
            content = await page.inner_text("body")
            lines = [l.strip() for l in content.split("\n") if l.strip()]
            for i, line in enumerate(lines):
                if line == "All" and i + 8 < len(lines):
                    data["total_contracts"] = lines[i + 2] if i + 2 < len(lines) else ""
                    data["oi_ch_24h"] = lines[i + 6] if i + 6 < len(lines) else ""
                    data["oi_24h_vol"] = lines[i + 7] if i + 7 < len(lines) else ""
                    print(f"  ✅ Fallback OI: {data['oi_ch_24h']}")
                    break
    except Exception as e:
        print(f"  ❌ OI error: {e}")

    return data


async def scrape_currencies(page, pair):
    url = URLS["currencies"].format(pair=pair)
    print(f"  Currencies: {url}")
    await page.goto(url, wait_until="domcontentloaded", timeout=30000)
    await asyncio.sleep(random.uniform(7, 12))

    data = {"long_short_liq": "", "price_24h": "", "net_flow": "", "ls_ratio": ""}

    try:
        content = await page.inner_text("body")
        lines = [l.strip() for l in content.split("\n") if l.strip()]

        for i, line in enumerate(lines):
            if "Price Performance" in line:
                for j in range(i + 1, min(i + 20, len(lines))):
                    if lines[j] == "24 hour" and j + 1 < len(lines):
                        val = lines[j + 1]
                        if "%" in val and len(val) < 15:
                            data["price_24h"] = val
                            print(f"  ✅ Price: {val}")
                            break
                break

        for i, line in enumerate(lines):
            if line == "24h Rekt":
                for j in range(i + 1, min(i + 15, len(lines))):
                    if "$" in lines[j] and ("M" in lines[j] or "K" in lines[j]):
                        if "/" in lines[j]:
                            data["long_short_liq"] = lines[j]
                            print(f"  ✅ Liq: {lines[j]}")
                            break

        for i, line in enumerate(lines):
            if "Long/Short Ratio" in line:
                for j in range(i + 1, min(i + 20, len(lines))):
                    if "Aggregated" in lines[j]:
                        k_start = j + 1
                        for k in range(k_start, min(k_start + 10, len(lines))):
                            if "Binance" in lines[k]:
                                bin_val = ""
                                okx_val = ""
                                bybit_val = ""
                                mexc_val = ""
                                
                                if k + 1 < len(lines):
                                    bin_val = lines[k + 1]
                                if "OKX" in str(lines[k + 2:k + 4]):
                                    for idx in range(k + 2, min(k + 5, len(lines))):
                                        if "OKX" in lines[idx] and idx + 1 < len(lines):
                                            okx_val = lines[idx + 1]
                                            break
                                if "Bybit" in str(lines[k + 2:k + 6]):
                                    for idx in range(k + 2, min(k + 7, len(lines))):
                                        if "Bybit" in lines[idx] and idx + 1 < len(lines):
                                            bybit_val = lines[idx + 1]
                                            break
                                if "Mexc" in str(lines[k + 2:k + 8]):
                                    for idx in range(k + 2, min(k + 9, len(lines))):
                                        if "Mexc" in lines[idx] and idx + 1 < len(lines):
                                            mexc_val = lines[idx + 1]
                                            break
                                
                                try:
                                    vals = []
                                    for v in [bin_val, okx_val, bybit_val, mexc_val]:
                                        if v and v.replace(".", "").replace(",", "").isdigit():
                                            vals.append(float(v))
                                    if vals:
                                        avg = sum(vals) / len(vals)
                                        data["ls_ratio"] = f"{bin_val}/{okx_val}/{bybit_val}/{mexc_val} (avg:{avg:.2f})"
                                        print(f"  ✅ LS Ratio: {data['ls_ratio']}")
                                except:
                                    pass
                                break
                        break

        for i, line in enumerate(lines):
            if "Net Flow" in line:
                for j in range(i + 1, min(i + 20, len(lines))):
                    if lines[j] == "24 hour" and j + 1 < len(lines):
                        val24 = lines[j + 1]
                    if lines[j] == "3 day" and j + 1 < len(lines):
                        val3d = lines[j + 1]
                        data["net_flow"] = f"{val24}/{val3d}"
                        print(f"  ✅ Net Flow: {data['net_flow']}")
                        break

    except Exception as e:
        print(f"  ❌ Currencies error: {e}")

    return data


async def scrape_volume(page, pair):
    url = URLS["volume"].format(pair=pair)
    print(f"  Volume: {url}")
    await page.goto(url, wait_until="domcontentloaded", timeout=30000)
    await asyncio.sleep(random.uniform(7, 12))

    data = {"vol_24h": "", "vol_7d": "", "vol_30d": "", "fut_spot_pct": ""}

    async def read_overview():
        try:
            content = await page.inner_text("body")
            lines = [l.strip() for l in content.split("\n") if l.strip()]
            
            for i, line in enumerate(lines):
                if "24 hour" in line and "$" in str(lines[i:i+5]):
                    for j in range(i, min(i + 5, len(lines))):
                        if "$" in lines[j] and ("B" in lines[j] or "M" in lines[j] or "K" in lines[j]):
                            data["vol_24h"] = parse_to_million(lines[j])
                            print(f"  ✅ Vol 24h: {data['vol_24h']}M")
                            break

            for i, line in enumerate(lines):
                if "24h Change" in line:
                    for j in range(i + 1, min(i + 15, len(lines))):
                        if "Futures" in lines[j] and "%" in str(lines[j:j+3]):
                            fut_pct = ""
                            spot_pct = ""
                            for k in range(j, min(j + 5, len(lines))):
                                if "%" in lines[k] and len(lines[k]) < 10:
                                    fut_pct = lines[k]
                                    break
                            if "Spot" in str(lines[j:j+5]):
                                for k in range(j, min(j + 8, len(lines))):
                                    if "Spot" in lines[k]:
                                        for m in range(k + 1, min(k + 4, len(lines))):
                                            if "%" in lines[m] and len(lines[m]) < 10:
                                                spot_pct = lines[m]
                                                break
                                        break
                            if fut_pct and spot_pct:
                                data["fut_spot_pct"] = f"{fut_pct}/{spot_pct}"
                                print(f"  ✅ Fut/Spot: {data['fut_spot_pct']}")
                            break
        except Exception as e:
            print(f"  ❌ Overview error: {e}")

    async def get_tab_vol(key, tab_text):
        try:
            await page.click(f"text={tab_text}")
            await asyncio.sleep(random.uniform(2, 4))
            
            content = await page.inner_text("body")
            lines = [l.strip() for l in content.split("\n") if l.strip()]
            
            in_card = False
            for i, line in enumerate(lines):
                if tab_text in line:
                    in_card = True
                if in_card and "Futures Volume" in line and i + 1 < len(lines):
                    for j in range(i + 1, min(i + 5, len(lines))):
                        val = lines[j]
                        if val.startswith("$") and ("B" in val or "M" in val or "K" in val):
                            data[key] = parse_to_million(val)
                            print(f"  ✅ {tab_text}: {data[key]}M")
                            return
                    break
        except Exception as e:
            print(f"  ❌ {tab_text} error: {e}")

    await read_overview()
    await get_tab_vol("vol_7d", "7 day")
    await get_tab_vol("vol_30d", "30 day")

    return data


def get_or_create_sheet(wb, pair):
    sheet_name = pair.capitalize()
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        for col_idx, col_name in enumerate(COLUMNS, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
    else:
        ws = wb[sheet_name]
    return ws


def find_next_row(ws):
    today = datetime.now().strftime("%Y-%m-%d")
    last_data_row = 1
    for row in range(2, ws.max_row + 1):
        date_val = ws.cell(row=row, column=1).value
        time_val = ws.cell(row=row, column=2).value
        if date_val is None and time_val is None:
            return row
        if date_val and today in str(date_val):
            if not time_val:
                return row
        if date_val:
            last_data_row = row
    return last_data_row + 1


def save_to_excel(pair, oi_data, curr_data, vol_data):
    print(f"\n  💾 Saving {pair}...")
    
    if os.path.exists(LOCK_FILE):
        try:
            lock_age = time.time() - os.path.getmtime(LOCK_FILE)
            if lock_age > 300:
                os.remove(LOCK_FILE)
                print(f"  🧹 Removed stale lock file")
        except:
            pass
    
    lock = FileLock(LOCK_FILE, timeout=120)
    
    try:
        with lock:
            if os.path.exists(EXCEL_PATH):
                wb = load_workbook(EXCEL_PATH)
                print(f"  📂 Opened existing file")
            else:
                wb = openpyxl.Workbook()
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]
                print(f"  📝 Created new file")

            ws = get_or_create_sheet(wb, pair)
            row = find_next_row(ws)
            now = datetime.now()

            v24 = vol_data.get("vol_24h", "")
            v7d = vol_data.get("vol_7d", "")
            v30d = vol_data.get("vol_30d", "")
            fut_spot_pct = vol_data.get("fut_spot_pct", "")

            ratio_7d, change_7d, text_7d = calc_vol_ratio(v24, v7d, 7)
            ratio_30d, change_30d, text_30d = calc_vol_ratio(v24, v30d, 30)

            def write_cell(col, value, is_negative=None):
                cell = ws.cell(row=row, column=col, value=value)
                if is_negative is True:
                    cell.fill = RED_FILL
                    cell.font = RED_FONT
                elif is_negative is False:
                    cell.fill = BLUE_FILL
                    cell.font = BLUE_FONT
                return cell

            oi_ch = oi_data.get("oi_ch_24h", "")
            price = curr_data.get("price_24h", "")
            net_flow = curr_data.get("net_flow", "")
            long_short_liq = curr_data.get("long_short_liq", "")
            
            liq_ratio = calculate_liquidation_ratio(long_short_liq)

            ws.cell(row=row, column=1, value=now.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=2, value=now.strftime("%I:%M:%S %p"))
            ws.cell(row=row, column=3, value=long_short_liq)
            
            if liq_ratio is not None:
                c4 = ws.cell(row=row, column=4, value=liq_ratio)
                if liq_ratio < -0.3:
                    c4.fill = RED_FILL
                    c4.font = RED_FONT
                elif liq_ratio > 0.3:
                    c4.fill = GREEN_FILL
                    c4.font = GREEN_FONT
            else:
                ws.cell(row=row, column=4, value="")
            
            ws.cell(row=row, column=5, value=oi_data.get("total_contracts", ""))
            write_cell(6, oi_ch, oi_ch.startswith("-") if oi_ch else None)
            ws.cell(row=row, column=7, value=oi_data.get("oi_24h_vol", ""))
            write_cell(8, fut_spot_pct, fut_spot_pct.startswith("-") if fut_spot_pct else None)
            write_cell(9, price, price.startswith("-") if price else None)

            ls_val = curr_data.get("ls_ratio", "")
            c10 = ws.cell(row=row, column=10, value=ls_val)
            try:
                avg_str = ls_val.split("avg:")[-1].strip() if "avg:" in ls_val else ls_val
                avg = float(avg_str.replace(")", ""))
                if avg > 1.0:
                    c10.fill = GREEN_FILL
                    c10.font = GREEN_FONT
                else:
                    c10.fill = RED_FILL
                    c10.font = RED_FONT
            except:
                pass

            ws.cell(row=row, column=11, value=v24)

            c12 = ws.cell(row=row, column=12, value=text_7d)
            if change_7d is not None:
                c12.fill = RED_FILL if change_7d < 0 else BLUE_FILL
                c12.font = RED_FONT if change_7d < 0 else BLUE_FONT

            c13 = ws.cell(row=row, column=13, value=text_30d)
            if change_30d is not None:
                c13.fill = RED_FILL if change_30d < 0 else BLUE_FILL
                c13.font = RED_FONT if change_30d < 0 else BLUE_FONT

            write_cell(14, net_flow, net_flow.startswith("-") if net_flow else None)

            wb.save(EXCEL_PATH)
            print(f"  ✅ Row {row}")

    except PermissionError:
        print(f"  ❌ Excel is open! Close it.")
    except Exception as e:
        import traceback
        print(f"  ❌ Excel error: {e}")
        traceback.print_exc()


async def collect_pair(browser, pair):
    print(f"\n{'='*50}")
    print(f"📊 {pair}")
    print(f"{'='*50}")

    context = await browser.new_context(
        user_agent=random.choice(USER_AGENTS),
        viewport={"width": random.randint(1280, 1920), "height": random.randint(720, 1080)},
        locale="en-US",
        timezone_id="America/New_York",
    )

    await context.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
        Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]});
        window.chrome = {runtime: {}};
        Object.defineProperty(navigator, 'languages', {get: () => ['en-US', 'en']});
    """)

    page = await context.new_page()

    try:
        oi_data = await scrape_open_interest(page, pair)
        await asyncio.sleep(random.uniform(3, 5))

        curr_data = await scrape_currencies(page, pair)
        await asyncio.sleep(random.uniform(3, 5))

        vol_data = await scrape_volume(page, pair)

        save_to_excel(pair, oi_data, curr_data, vol_data)

    except Exception as e:
        print(f"❌ Error: {e}")
    finally:
        await context.close()


async def main():
    print(f"\n🚀 Coinglass Data Collector - PART 4")
    print(f"⏰ Time: {datetime.now().strftime('%Y-%m-%d %I:%M:%S %p')}")
    print(f"📈 Pairs: {len(PAIRS)} (Part 4 of 4)")
    print(f"⏱️  Estimated time: ~15 minutes\n")

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            slow_mo=500,
        )

        for idx, pair in enumerate(PAIRS, 1):
            print(f"\n[{idx}/{len(PAIRS)}]")
            await collect_pair(browser, pair)
            
            delay = random.uniform(15, 30)
            print(f"⏳ Wait {delay:.0f}s...")
            await asyncio.sleep(delay)

        await browser.close()

    print(f"\n✅ Part 4 Done! Saved to: {EXCEL_PATH}")


if __name__ == "__main__":
    asyncio.run(main())
